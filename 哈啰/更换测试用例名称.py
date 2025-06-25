import datetime

import openpyxl


wb = openpyxl.load_workbook('测试用例.xlsx')
# 获取工作表的名字
print(wb.sheetnames)
# 获取工作表 ---> Worksheet
sheet = wb.worksheets[0]
# 获得单元格的范围
print(sheet.dimensions)
# 获得行数和列数
print(sheet.max_row, sheet.max_column) 

# 

# 获取第二列的值
for row_ch in range(2, sheet.max_row + 1):
    cell_value = sheet.cell(row=row_ch, column=2).value
    if cell_value and isinstance(cell_value, str) and '[线路]' in cell_value:
        # 提取 [线路] 后面的值
        line_value = cell_value.split('[线路]')[-1].strip()
        # 获取第三列单元格
        third_col_cell = sheet.cell(row=row_ch, column=3)
        # 获取第三列原有的值，若为空则设为空字符串
        original_value = third_col_cell.value if third_col_cell.value else ''
        # 将提取的值添加到第三列原有值前面
        new_value = line_value + " " + str(original_value)
        third_col_cell.value = new_value

# 保存修改后的工作簿
wb.save('测试用例.xlsx')